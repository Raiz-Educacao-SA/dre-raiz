"""
Gera Excel dinâmico com dados da tabela transactions do Supabase.
Uso: python scripts/gerar_excel_transactions.py
     python scripts/gerar_excel_transactions.py --output "C:/caminho/arquivo.xlsx"

O arquivo gerado inclui:
  - Aba "Transactions" com todos os dados formatados como Tabela Excel
  - Aba "PowerQuery" com o código M para configurar refresh nativo no Excel
"""

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import sys
import os
from datetime import datetime

# ─── Configuração Supabase ───────────────────────────────────────────────────
SUPABASE_URL = "https://vafmufhlompwsdrlhkfz.supabase.co"
SERVICE_ROLE_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InZhZm11Zmhsb21wd3Nkcmxoa2Z6Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc2OTQzMjI5MSwiZXhwIjoyMDg1MDA4MjkxfQ"
    ".m0viLArNl57ExdNlRoEuNNZH0n_0iKSaa81Fyj2ekpA"
)

# ─── Definição de Colunas ────────────────────────────────────────────────────
COLUMNS = [
    ("Chave ID",         "chave_id",       15,  "text"),
    ("Data",             "date",           12,  "text"),
    ("Descrição",        "description",    40,  "text"),
    ("Conta Contábil",   "conta_contabil", 18,  "text"),
    ("Valor",            "amount",         16,  "number"),
    ("Tipo",             "type",           16,  "text"),
    ("Cenário",          "scenario",       12,  "text"),
    ("Status",           "status",         12,  "text"),
    ("Marca",            "marca",          8,   "text"),
    ("Filial",           "filial",         10,  "text"),
    ("Nome Filial",      "nome_filial",    25,  "text"),
    ("Tag0",             "tag0",           25,  "text"),
    ("Tag01",            "tag01",          25,  "text"),
    ("Tag02",            "tag02",          25,  "text"),
    ("Tag03",            "tag03",          25,  "text"),
    ("Recorrente",       "recurring",      12,  "text"),
    ("Ticket",           "ticket",         12,  "text"),
    ("Fornecedor",       "vendor",         25,  "text"),
    ("Nat. Orçamentária","nat_orc",        18,  "text"),
]


def fetch_all_transactions():
    """Busca todos os registros da tabela transactions via REST API (paginado)."""
    all_records = []
    limit = 1000
    offset = 0
    headers = {
        "apikey": SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SERVICE_ROLE_KEY}",
        "Prefer": "count=exact",
    }

    print("Buscando dados do Supabase...")
    while True:
        url = (
            f"{SUPABASE_URL}/rest/v1/transactions"
            f"?select=*&order=date.desc,marca,filial"
            f"&limit={limit}&offset={offset}"
        )
        resp = requests.get(url, headers=headers)
        # 200 = todos os dados, 206 = partial content (há mais páginas)
        if resp.status_code not in (200, 206):
            print(f"ERRO: HTTP {resp.status_code} — {resp.text[:200]}")
            sys.exit(1)

        data = resp.json()
        if not data:
            break

        all_records.extend(data)
        print(f"  ... {len(all_records)} registros carregados")

        # 200 = última página ou tudo de uma vez
        if resp.status_code == 200 or len(data) < limit:
            break
        offset += limit

    print(f"Total: {len(all_records)} registros\n")
    return all_records


def create_excel(transactions, output_path):
    """Cria workbook Excel com dados formatados como Tabela."""
    wb = openpyxl.Workbook()

    # ── Aba 1: Transactions ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Transactions"

    # Estilos
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    number_fmt = '#,##0.00'
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Headers
    for col_idx, (header, _, width, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Dados
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    for row_idx, tx in enumerate(transactions, 2):
        for col_idx, (_, key, _, col_type) in enumerate(COLUMNS, 1):
            raw = tx.get(key)
            if col_type == "number" and raw is not None:
                try:
                    value = float(raw)
                except (ValueError, TypeError):
                    value = raw
            else:
                value = raw if raw is not None else ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            if col_type == "number" and isinstance(value, (int, float)):
                cell.number_format = number_fmt
                cell.alignment = Alignment(horizontal="right")

        # Zebra stripes
        if row_idx % 2 == 0:
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = alt_fill

    # Criar Tabela Excel formal (permite right-click → Refresh com Power Query)
    if len(transactions) > 0:
        last_col = get_column_letter(len(COLUMNS))
        last_row = len(transactions) + 1
        table_ref = f"A1:{last_col}{last_row}"
        table = Table(displayName="TabelaTransactions", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    # Freeze header
    ws.freeze_panes = "A2"
    # Auto-filter já vem com a Table

    # ── Aba 2: Instruções Power Query ────────────────────────────────────────
    ws2 = wb.create_sheet("PowerQuery - Instruções")
    ws2.sheet_properties.tabColor = "4472C4"

    instructions = [
        ("COMO CONFIGURAR REFRESH AUTOMÁTICO (Power Query)", "", ""),
        ("", "", ""),
        ("Com Power Query, você pode clicar com botão direito na tabela → Atualizar", "", ""),
        ("e o Excel busca os dados novos do Supabase automaticamente.", "", ""),
        ("", "", ""),
        ("PASSO A PASSO:", "", ""),
        ("1. Aba 'Dados' → 'Obter Dados' → 'De Outras Fontes' → 'Consulta Nula'", "", ""),
        ("2. No Editor do Power Query, clique em 'Editor Avançado'", "", ""),
        ("3. Apague tudo e cole o código M da coluna C abaixo", "", ""),
        ("4. Clique 'Concluído' → 'Fechar e Carregar'", "", ""),
        ("5. Pronto! Agora clique com botão direito → 'Atualizar' a qualquer momento", "", ""),
        ("", "", ""),
        ("CÓDIGO M (copie da célula C13):", "", ""),
    ]

    m_code = f'''let
    Url = "{SUPABASE_URL}/rest/v1/transactions?select=*&order=date.desc,marca,filial&limit=100000",
    Headers = [
        apikey = "{SERVICE_ROLE_KEY}",
        Authorization = "Bearer {SERVICE_ROLE_KEY}"
    ],
    Source = Json.Document(Web.Contents(Url, [Headers=Headers])),
    Table = Table.FromList(Source, Splitter.SplitByNothing(), null, null, ExtraValues.Error),
    Expanded = Table.ExpandRecordColumn(Table, "Column1", {{
        "chave_id", "date", "description", "conta_contabil", "amount",
        "type", "scenario", "status", "marca", "filial", "nome_filial",
        "tag0", "tag01", "tag02", "tag03", "recurring", "ticket", "vendor", "nat_orc"
    }}, {{
        "Chave ID", "Data", "Descrição", "Conta Contábil", "Valor",
        "Tipo", "Cenário", "Status", "Marca", "Filial", "Nome Filial",
        "Tag0", "Tag01", "Tag02", "Tag03", "Recorrente", "Ticket", "Fornecedor", "Nat. Orçamentária"
    }}),
    Typed = Table.TransformColumnTypes(Expanded, {{
        {{"Valor", type number}},
        {{"Data", type text}}
    }})
in
    Typed'''

    title_font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    step_font = Font(name="Calibri", size=11)
    code_font = Font(name="Consolas", size=10, color="2E4057")
    code_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    for row_idx, (a, b, c) in enumerate(instructions, 1):
        cell_a = ws2.cell(row=row_idx, column=1, value=a)
        if row_idx == 1:
            cell_a.font = title_font
        elif a.startswith(("1.", "2.", "3.", "4.", "5.")):
            cell_a.font = Font(name="Calibri", size=11, bold=True, color="2E75B6")
        else:
            cell_a.font = step_font

    # Código M na célula C13
    m_cell = ws2.cell(row=13, column=3, value=m_code)
    m_cell.font = code_font
    m_cell.fill = code_fill
    m_cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws2.column_dimensions["A"].width = 75
    ws2.column_dimensions["C"].width = 100
    ws2.merge_cells("A1:B1")

    # ── Aba 3: Alternativa — script Python ───────────────────────────────────
    ws3 = wb.create_sheet("Script - Alternativa")
    ws3.sheet_properties.tabColor = "70AD47"
    alt_instructions = [
        "ALTERNATIVA: Atualizar via script Python",
        "",
        "Se preferir não usar Power Query, você pode re-executar o script:",
        "",
        '  Opção 1: Duplo-clique no arquivo "atualizar_excel_transactions.bat"',
        "  Opção 2: No terminal, rode: python scripts/gerar_excel_transactions.py",
        "",
        "O script busca todos os dados do Supabase e gera este Excel novamente.",
        f"Última atualização: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
    ]
    for row_idx, text in enumerate(alt_instructions, 1):
        cell = ws3.cell(row=row_idx, column=1, value=text)
        if row_idx == 1:
            cell.font = Font(name="Calibri", bold=True, size=14, color="70AD47")
        else:
            cell.font = Font(name="Calibri", size=11)
    ws3.column_dimensions["A"].width = 80

    # ── Salvar ───────────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"Excel salvo em: {output_path}")
    print(f"  - {len(transactions)} registros na aba 'Transactions'")
    print(f"  - Código Power Query na aba 'PowerQuery - Instruções'")
    print(f"  - Instruções do script na aba 'Script - Alternativa'")


def main():
    # Output path
    output = None
    if "--output" in sys.argv:
        idx = sys.argv.index("--output")
        if idx + 1 < len(sys.argv):
            output = sys.argv[idx + 1]

    if not output:
        # Default: mesma pasta do projeto
        script_dir = os.path.dirname(os.path.abspath(__file__))
        project_dir = os.path.dirname(script_dir)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output = os.path.join(project_dir, f"Transactions_Supabase_{timestamp}.xlsx")

    print("=" * 60)
    print("  Gerador de Excel — Transactions Supabase")
    print("=" * 60)
    print()

    transactions = fetch_all_transactions()

    if not transactions:
        print("AVISO: Nenhum registro encontrado. Verifique a conexão.")
        sys.exit(1)

    create_excel(transactions, output)

    print()
    print("Pronto! Abra o Excel e veja as instruções na aba 'PowerQuery'")
    print("para configurar refresh automático com botão direito.")


if __name__ == "__main__":
    main()
